import csv
import json
import datetime
import calendar
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.views import LoginView
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.models import User
from django.db.models import Count, Q, Avg
from .forms import UserRegistrationForm, GoalForm, HabitForm, TaskForm, ReflectionForm, UserProfileForm
from .models import Goal, Task, Habit, HabitCompletion, Reflection, UserProfile

def register(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.set_password(form.cleaned_data['password'])
            user.save()  # Signal creates UserProfile automatically
            
            profile = user.userprofile
            age = form.cleaned_data.get('age')
            if age is not None:
                profile.age = age
                profile.save()
                
            login(request, user)
            return redirect('dashboard')
    else:
        form = UserRegistrationForm()
    return render(request, 'register.html', {'form': form})

class CustomLoginView(LoginView):
    template_name = 'login.html'

    def get_success_url(self):
        # Redirect admins to the admin dashboard, others to regular dashboard
        if self.request.user.is_superuser:
            return '/admin-panel/overview/'
        return super().get_success_url()

def get_task_score(user, target_date):
    # Only evaluate tasks that were due on the target date to measure daily hit rate properly
    # This prevents the score from being a lifetime average.
    tasks = Task.objects.filter(user=user, due_date=target_date)
    if not tasks.exists():
        return 0
    total_possible = tasks.count() * 100
    actual_score = 0
    for task in tasks:
        if task.status == 'Completed':
            actual_score += 100
        elif task.task_type == 'trackable':
            actual_score += task.progress
    return (actual_score / total_possible) * 100

def get_habit_score(user, target_date):
    from datetime import date
    habits = Habit.objects.filter(user=user)
    if not habits.exists():
        return 0
    total = 0
    for habit in habits:
        completion = HabitCompletion.objects.filter(
            habit=habit,
            date=target_date
        ).first()
        total += (
            completion.completion_percentage
            if completion else 0
        )
    return total / habits.count()

def get_daily_score(user, target_date):
    habit_score = get_habit_score(user, target_date)
    task_score = get_task_score(user, target_date)
    return round((habit_score + task_score) / 2)

@login_required
def dashboard(request):
    user = request.user
    today = timezone.localdate()
    current_hour = timezone.localtime().hour
    
    # --- Greeting Logic ---
    if 5 <= current_hour < 12:
        greeting_time = "Good Morning"
    elif 12 <= current_hour < 17:
        greeting_time = "Good Afternoon"
    else:
        greeting_time = "Good Evening"

    try:
        display_name = user.userprofile.full_name.split()[0] if user.userprofile.full_name else user.username
    except UserProfile.DoesNotExist:
        display_name = user.username
        
    quotes = [
        "Focus on being productive instead of busy.",
        "Small steps every day lead to big results.",
        "Your future is created by what you do today, not tomorrow.",
        "Don't count the days, make the days count.",
        "Success is sum of small efforts repeated day in and day out.",
        "Action is the foundational key to all success.",
        "The secret of getting ahead is getting started."
    ]
    daily_quote = quotes[today.toordinal() % len(quotes)]
    
    # --- Quick Stats ---
    total_tasks_due_today = Task.objects.filter(user=user, start_date__lte=today, status='Pending').count()
    total_active_goals = Goal.objects.filter(user=user, status='Active').count()
    habits_logged_today = HabitCompletion.objects.filter(habit__user=user, date=today).values('habit').distinct().count()
    
    # --- Today's Tasks (Sorted) ---
    raw_today_tasks = list(Task.objects.filter(user=user, start_date__lte=today, status='Pending'))
    def sort_key(t):
        if t.due_date:
            if t.due_date < today: return 0  # Overdue
            elif t.due_date == today: return 1  # Due today
            else: return 2  # Future
        return 3  # No due date
        
    today_tasks = sorted(raw_today_tasks, key=sort_key)
    
    for t in today_tasks:
        if t.due_date:
            delta = (today - t.due_date).days
            if delta > 0:
                t.date_label = f"{delta} day{'s' if delta > 1 else ''} overdue"
                t.ui_status = "overdue"
            elif delta == 0:
                t.date_label = "Due today"
                t.ui_status = "today"
            else:
                t.date_label = f"Due in {abs(delta)} day{'s' if abs(delta) > 1 else ''}"
                t.ui_status = "future"
        else:
            t.date_label = "No due date"
            t.ui_status = "none"
            
    # --- Goal Progress ---
    active_goals = Goal.objects.filter(user=user, status='Active')
    for g in active_goals:
        # Use the model property which correctly handles both task-based and habit-based goals
        g.progress_pct = g.progress_percentage

        if g.goal_type == 'habit_based':
            # For habit-based goals show habit count instead of tasks
            g.completed_count = round(g.progress_percentage)
            g.total_count = 100  # percentage out of 100
        else:
            total = g.tasks.count()
            completed = g.tasks.filter(status='Completed').count()
            g.completed_count = completed
            g.total_count = total

        if g.progress_pct <= 30: g.progress_color = 'danger'
        elif g.progress_pct <= 70: g.progress_color = 'warning'
        else: g.progress_color = 'success'

    # --- Today's Habits ---
    today_habits = Habit.objects.filter(user=user)
    for h in today_habits:
        completion = HabitCompletion.objects.filter(habit=h, date=today).first()
        h.completion_today = completion.completion_percentage if completion else None
        h.streak = calculate_streak(h)
        
        pct = h.completion_today or 0
        if pct == 0: h.progress_color = 'secondary'
        elif pct <= 29: h.progress_color = 'danger'
        elif pct <= 69: h.progress_color = 'warning'
        else: h.progress_color = 'success'
        
    # --- Productivity Graph ---
    last_7_days = [(today - datetime.timedelta(days=i)) for i in range(6, -1, -1)]
    chart_labels = [d.strftime('%a') for d in last_7_days]
    chart_data = []
    
    all_zeroes = True
    for d in last_7_days:
        productivity_score = get_daily_score(user, d)
        chart_data.append(productivity_score)
        if productivity_score > 0:
            all_zeroes = False

    has_activity_data = not all_zeroes
    avg_score = round(sum(chart_data) / len(chart_data)) if chart_data else 0
    
    context = {
        'greeting_time': greeting_time,
        'display_name': display_name,
        'today_date': today,
        'daily_quote': daily_quote,
        'total_tasks_due_today': total_tasks_due_today,
        'total_active_goals': total_active_goals,
        'habits_logged_today': habits_logged_today,
        'today_tasks': today_tasks,
        'active_goals': active_goals,
        'today_habits': today_habits,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'has_activity_data': has_activity_data,
        'avg_score': avg_score,
    }
    
    return render(request, 'dashboard.html', context)

# ==================== PROFILE ====================
@login_required
def profile_view(request):
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=profile, user_obj=request.user)
        if form.is_valid():
            form.save()
            return redirect('profile_view')
    else:
        form = UserProfileForm(instance=profile, user_obj=request.user)
        
    return render(request, 'profile.html', {'form': form, 'profile': profile})

# ==================== GOALS ====================
# Helper: check and trigger goal complete prompt
def check_goal_completion(goal):
    if goal and goal.progress_percentage == 100:
        return True
    return False

# Goals list view
@login_required
def goals_list(request):
    goals = Goal.objects.filter(
        user=request.user
    ).order_by('-created_at')

    # Attach progress and completion flag
    for goal in goals:
        goal.progress = goal.progress_percentage
        goal.show_complete_prompt = (
            goal.progress == 100 and
            goal.status == 'Active'
        )

    context = {
        'goals': goals,
        'active_goals': goals.filter(status='Active'),
        'completed_goals': goals.filter(
            status='Completed'
        ),
    }
    return render(request, 'goals.html', context)

# Create goal view
@login_required
def create_goal(request):
    if request.method == 'POST':
        form = GoalForm(request.POST)
        if form.is_valid():
            goal = form.save(commit=False)
            goal.user = request.user
            goal.save()
            return redirect('goals_list')
    else:
        form = GoalForm()
    return render(
        request,
        'goal_form.html',
        {'form': form}
    )

# Edit goal view
@login_required
def edit_goal(request, goal_id):
    goal = get_object_or_404(
        Goal,
        id=goal_id,
        user=request.user
    )
    if request.method == 'POST':
        form = GoalForm(request.POST, instance=goal)
        if form.is_valid():
            form.save()
            return redirect('goals_list')
    else:
        form = GoalForm(instance=goal)
    return render(
        request,
        'goal_form.html',
        {'form': form, 'goal': goal}
    )

# Delete goal view
@login_required
def delete_goal(request, goal_id):
    goal = get_object_or_404(
        Goal,
        id=goal_id,
        user=request.user
    )
    if request.method == 'POST':
        goal.delete()
        return redirect('goals_list')
    return render(
        request,
        'confirm_delete.html',
        {'object': goal}
    )

# Complete goal view
@login_required
def complete_goal(request, goal_id):
    from django.contrib import messages
    goal = get_object_or_404(
        Goal,
        id=goal_id,
        user=request.user
    )
    # Only allow complete at 100%
    if goal.progress_percentage == 100:
        from django.utils import timezone
        goal.status = 'Completed'
        goal.completed_at = timezone.now()
        goal.save()
        messages.success(request, f"Goal '{goal.title}' successfully completed!")
    else:
        messages.error(request, f"Goal '{goal.title}' cannot be completed until progress is 100%. User attempted to bypass.")
    return redirect('goals_list')


# ==================== HABITS ====================
# Updated streak calculation
def calculate_streak(habit):
    from datetime import date, timedelta
    today = date.today()
    streak = 0
    current_date = today

    while True:
        completion = HabitCompletion.objects.filter(
            habit=habit,
            date=current_date
        ).first()

        if habit.habit_type == 'checkbox':
            completed = (
                completion and
                completion.completion_percentage == 100
            )
        else:
            completed = (
                completion and
                completion.completion_percentage > 0
            )

        if completed:
            streak += 1
            current_date -= timedelta(days=1)
        else:
            break

    return streak

# Updated weekly data
def get_weekly_data(habit):
    from datetime import date, timedelta
    today = date.today()
    weekly_data = []

    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        completion = HabitCompletion.objects.filter(
            habit=habit,
            date=day
        ).first()

        percentage = (
            completion.completion_percentage
            if completion else 0
        )
        logged = completion is not None

        weekly_data.append({
            'day': day.strftime('%a'),
            'date': day,
            'percentage': percentage,
            'logged': logged
        })

    return weekly_data

@login_required
def habits_list(request):
    user = request.user
    habits = Habit.objects.filter(user=user).order_by('-created_at')
    today = timezone.localdate()
    
    user_goals = Goal.objects.filter(user=user, status='Active')
    
    for habit in habits:
        completion = HabitCompletion.objects.filter(habit=habit, date=today).first()
        habit.completion_percentage_today = completion.completion_percentage if completion else 0
        habit.current_streak = calculate_streak(habit)
        habit.weekly_data = get_weekly_data(habit)

    if request.method == 'POST':
        form = HabitForm(request.POST)
        if form.is_valid():
            h = form.save(commit=False)
            h.user = user
            h.save()
            return redirect('habits_list')
    else:
        form = HabitForm()
        
    context = {
        'habits': habits,
        'user_goals': user_goals,
        'form': form,
        'today_completion': 0 # template fallback if needed
    }
    return render(request, 'habits.html', context)

@login_required
def edit_habit(request, habit_id):
    habit = get_object_or_404(Habit, id=habit_id, user=request.user)
    if request.method == 'POST':
        form = HabitForm(request.POST, instance=habit)
        if form.is_valid():
            form.save()
            return redirect('habits_list')
    return redirect('habits_list')

@login_required
def delete_habit(request, habit_id):
    habit = get_object_or_404(Habit, id=habit_id, user=request.user)
    if request.method == 'POST':
        habit.delete()
    return redirect('habits_list')

@login_required
def complete_habit(request, habit_id):
    return save_habit_log(request, habit_id)

# Save habit log view
@login_required
def save_habit_log(request, habit_id):
    if request.method == 'POST':
        habit = get_object_or_404(
            Habit,
            id=habit_id,
            user=request.user
        )
        today = timezone.localdate()

        if habit.habit_type == 'checkbox':
            is_done = request.POST.get(
                'is_done'
            ) == 'true'
            percentage = 100 if is_done else 0

        elif habit.habit_type == 'slider':
            percentage = int(
                request.POST.get('percentage', 0)
            )
            percentage = max(0, min(100, percentage))
        else:
            percentage = 0

        HabitCompletion.objects.update_or_create(
            habit=habit,
            date=today,
            defaults={
                'completion_percentage': percentage
            }
        )
        return redirect('habits_list')

# ==================== TASKS ====================

@login_required
def tasks_list(request):
    user = request.user
    today = timezone.localdate()
    all_tasks = Task.objects.filter(user=user)
    
    # 1. Overdue
    overdue_tasks = all_tasks.filter(due_date__lt=today, status='Pending').order_by('due_date')
    
    # 2. Today
    today_tasks = all_tasks.filter(start_date__lte=today, status='Pending', due_date=today).order_by('created_at')
    
    # 3. Pending
    raw_pending = all_tasks.filter(status='Pending').exclude(id__in=overdue_tasks).exclude(id__in=today_tasks)
    pending_tasks = sorted(raw_pending, key=lambda t: (t.due_date is None, t.due_date))
    
    # 4. Completed
    completed_tasks = all_tasks.filter(status='Completed')
    
    # Process Labels
    all_processed = list(overdue_tasks) + list(today_tasks) + list(pending_tasks) + list(completed_tasks)
    for task in all_processed:
        if task.due_date:
            delta = (task.due_date - today).days
            if delta < 0:
                task.due_label = f"{abs(delta)} day{'s' if abs(delta) > 1 else ''} overdue"
                task.due_color = 'danger'
            elif delta == 0:
                task.due_label = "Due today"
                task.due_color = 'warning'
            else:
                task.due_label = f"Due in {delta} day{'s' if delta > 1 else ''}"
                task.due_color = 'text-primary'
        else:
            task.due_label = "No due date"
            task.due_color = 'secondary'
            
    # For Dropdowns
    user_goals = Goal.objects.filter(user=user, status='Active')
    for g in user_goals:
        total = g.tasks.count()
        completed = g.tasks.filter(status='Completed').count()
        g.is_100_percent = (total > 0 and total == completed)
        
    user_habits = Habit.objects.filter(user=user)

    if request.method == 'POST':
        form = TaskForm(request.POST)
        if form.is_valid():
            t = form.save(commit=False)
            t.user = user
            t.save()
            return redirect('tasks_list')
    else:
        form = TaskForm()
        
    context = {
        'overdue_tasks': overdue_tasks,
        'today_tasks': today_tasks,
        'pending_tasks': pending_tasks,
        'completed_tasks': completed_tasks,
        'user_goals': user_goals,
        'user_habits': user_habits,
        'form': form
    }
    return render(request, 'tasks.html', context)

# Save task progress view
@login_required
def save_task_progress(request, task_id):
    if request.method == 'POST':
        task = get_object_or_404(
            Task,
            id=task_id,
            user=request.user,
            task_type='trackable'
        )
        progress = int(
            request.POST.get('progress', 0)
        )
        progress = max(0, min(100, progress))
        task.progress = progress
        task.save()

        # Check goal completion
        goal_complete = False
        goal_title = ''
        goal_id = None
        if task.goal:
            if task.goal.progress_percentage == 100:
                goal_complete = True
                goal_title = task.goal.title
                goal_id = task.goal.id

        return redirect('tasks_list')

# Complete task view
@login_required
def complete_task(request, task_id):
    task = get_object_or_404(
        Task,
        id=task_id,
        user=request.user
    )
    task.status = 'Completed'
    task.progress = 100
    task.save()
    return redirect('tasks_list')

# Delete task view
@login_required
def delete_task(request, task_id):
    task = get_object_or_404(
        Task,
        id=task_id,
        user=request.user
    )
    if request.method == 'POST':
        task.delete()
        return redirect('tasks_list')
    return render(
        request,
        'confirm_delete.html',
        {'object': task}
    )

@login_required
def edit_task(request, task_id):
    task = get_object_or_404(Task, id=task_id, user=request.user)
    if request.method == 'POST':
        form = TaskForm(request.POST, instance=task)
        if form.is_valid():
            form.save()
            return redirect('tasks_list')
    else:
        form = TaskForm(instance=task)
    return render(request, 'edit_task.html', {'form': form, 'task': task})

# ==================== REFLECTIONS ====================
def calculate_reflection_streak(user):
    today = timezone.localdate()
    streak = 0
    current_date = today

    while True:
        exists = Reflection.objects.filter(
            user=user,
            date=current_date
        ).exists()

        if exists:
            streak += 1
            current_date -= datetime.timedelta(days=1)
        else:
            break

    return streak

@login_required
def reflections_list(request):
    user = request.user
    today = timezone.localdate()
    
    # Handle Save
    if request.method == 'POST':
        wins = request.POST.get('wins', '')
        challenges = request.POST.get('challenges', '')
        tomorrow = request.POST.get('tomorrow', '')
        
        # At least one field must be filled
        if wins or challenges or tomorrow:
            content = {
                'wins': wins,
                'challenges': challenges,
                'tomorrow': tomorrow
            }
            Reflection.objects.update_or_create(
                user=user,
                date=today,
                defaults={'content': content}
            )
        return redirect('reflections_list')

    # Get Calendar Query Context
    try:
        current_year = int(request.GET.get('year', today.year))
        current_month = int(request.GET.get('month', today.month))
    except ValueError:
        current_year = today.year
        current_month = today.month

    # Bounding checks
    user_joined_date = user.date_joined.date()
    
    # Prevent future navigation
    if current_year > today.year or (current_year == today.year and current_month > today.month):
        current_year = today.year
        current_month = today.month
        
    # Generate Calendar Grid
    raw_cal = calendar.monthcalendar(current_year, current_month)
    
    # Get Reflections for this month
    monthly_reflections = Reflection.objects.filter(
        user=user,
        date__year=current_year,
        date__month=current_month
    )
    
    # Pack for Frontend
    reflection_dates_str = [r.date.strftime('%Y-%m-%d') for r in monthly_reflections]
    
    cal_data = []
    for week in raw_cal:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'day': 0, 'date_str': ''})
            else:
                date_str = f"{current_year}-{current_month:02d}-{day:02d}"
                week_data.append({
                    'day': day,
                    'date_str': date_str,
                    'has_reflection': date_str in reflection_dates_str,
                    'is_today': date_str == today.strftime('%Y-%m-%d')
                })
        cal_data.append(week_data)
    
    reflections_dict = {}
    for r in monthly_reflections:
        reflections_dict[r.date.strftime('%Y-%m-%d')] = {
            'id': r.id,
            'wins': r.content.get('wins', ''),
            'challenges': r.content.get('challenges', ''),
            'tomorrow': r.content.get('tomorrow', '')
        }
        
    # Navigation Math
    prev_month = current_month - 1
    prev_year = current_year
    if prev_month == 0:
        prev_month = 12
        prev_year -= 1
        
    next_month = current_month + 1
    next_year = current_year
    if next_month == 13:
        next_month = 1
        next_year += 1
        
    # Disable Previous if goes before user Registration
    disable_prev = False
    if prev_year < user_joined_date.year or (prev_year == user_joined_date.year and prev_month < user_joined_date.month):
        disable_prev = True
        
    disable_next = False
    if next_year > today.year or (next_year == today.year and next_month > today.month):
        disable_next = True
        
    streak = calculate_reflection_streak(user)
    total_empty = Reflection.objects.filter(user=user).count() == 0

    # Get today's reflection ID if it exists (for the cancel button)
    today_reflection = Reflection.objects.filter(user=user, date=today).first()
    today_reflection_id = today_reflection.id if today_reflection else None

    context = {
        'today_str': today.strftime('%Y-%m-%d'),
        'current_month_name': calendar.month_name[current_month],
        'current_year': current_year,
        'current_month': current_month,
        'cal': cal_data,
        'reflection_dates_str': reflection_dates_str,
        'reflections_json': json.dumps(reflections_dict),
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
        'disable_prev': disable_prev,
        'disable_next': disable_next,
        'streak': streak,
        'total_empty': total_empty,
        'today_reflection_id': today_reflection_id,
    }
    
    return render(request, 'reflections.html', context)

@login_required
def delete_reflection(request, reflection_id):
    reflection = get_object_or_404(Reflection, id=reflection_id, user=request.user)
    if request.method == 'POST':
        reflection.delete()
    return redirect('reflections_list')

@login_required
def edit_reflection(request, reflection_id):
    reflection = get_object_or_404(Reflection, id=reflection_id, user=request.user)
    if request.method == 'POST':
        form = ReflectionForm(request.POST, instance=reflection)
        if form.is_valid():
            form.save()
            return redirect('reflections_list')
    else:
        form = ReflectionForm(instance=reflection)
    return render(request, 'edit_reflection.html', {'form': form, 'reflection': reflection})

# ==================== REPORTS ====================

def get_report_summary(user, from_date, to_date):
    summary = {}
    summary['tasks'] = Task.objects.filter(
        user=user, created_at__date__gte=from_date, created_at__date__lte=to_date
    ).count()
    summary['habits'] = HabitCompletion.objects.filter(
        habit__user=user, date__gte=from_date, date__lte=to_date
    ).count()
    summary['goals'] = Goal.objects.filter(
        user=user, created_at__date__gte=from_date, created_at__date__lte=to_date
    ).count()
    summary['reflections'] = Reflection.objects.filter(
        user=user, date__gte=from_date, date__lte=to_date
    ).count()
    summary['total'] = sum(summary.values())
    return summary

@login_required
def report_summary_api(request):
    try:
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')
        if not from_date_str or not to_date_str:
            return JsonResponse({'error': 'Missing dates'}, status=400)
            
        from_date = datetime.datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date = datetime.datetime.strptime(to_date_str, '%Y-%m-%d').date()
        summary = get_report_summary(request.user, from_date, to_date)
        return JsonResponse(summary)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def reports_page(request):
    min_date = request.user.date_joined.date().strftime('%Y-%m-%d')
    today = datetime.date.today().strftime('%Y-%m-%d')
    return render(request, 'reports.html', {'min_date': min_date, 'today': today})

@login_required
def generate_report(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    include = request.GET.getlist('include')

    # Check if any data exists before generating report
    has_data = False
    if 'tasks' in include and Task.objects.filter(
        user=request.user, created_at__date__gte=from_date, created_at__date__lte=to_date
    ).exists():
        has_data = True
    if 'habits' in include and HabitCompletion.objects.filter(
        habit__user=request.user, date__gte=from_date, date__lte=to_date
    ).exists():
        has_data = True
    if 'goals' in include and Goal.objects.filter(
        user=request.user, created_at__date__gte=from_date, created_at__date__lte=to_date
    ).exists():
        has_data = True
    if 'reflections' in include and Reflection.objects.filter(
        user=request.user, date__gte=from_date, date__lte=to_date
    ).exists():
        has_data = True

    if not has_data:
        from django.http import JsonResponse
        return JsonResponse({'error': 'No data found for the selected date range.'}, status=404)

    # Helper to style header rows
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1C2340')

    def style_header(ws, headers):
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    # TASKS SHEET
    if 'tasks' in include:
        ws = wb.create_sheet('Tasks')
        style_header(ws, ['Title', 'Description', 'Status', 'Start Date', 'Due Date', 'Linked Goal', 'Created At'])
        tasks = Task.objects.filter(
            user=request.user, created_at__date__gte=from_date, created_at__date__lte=to_date
        )
        for task in tasks:
            ws.append([
                task.title,
                task.description or '',
                task.status,
                str(task.start_date) if task.start_date else '',
                str(task.due_date) if task.due_date else '',
                task.goal.title if task.goal else '',
                task.created_at.strftime('%Y-%m-%d')
            ])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].auto_size = True

    # HABITS SHEET
    if 'habits' in include:
        ws = wb.create_sheet('Habits')
        style_header(ws, ['Habit Name', 'Linked Goal', 'Date', 'Completion %'])
        completions = HabitCompletion.objects.filter(
            habit__user=request.user, date__gte=from_date, date__lte=to_date
        ).select_related('habit', 'habit__goal')
        for c in completions:
            ws.append([
                c.habit.habit_name,
                c.habit.goal.title if c.habit.goal else '',
                str(c.date),
                c.completion_percentage
            ])

    # GOALS SHEET
    if 'goals' in include:
        ws = wb.create_sheet('Goals')
        style_header(ws, ['Title', 'Description', 'Status', 'Progress %', 'Tasks Linked', 'Habits Linked', 'Created At'])
        goals = Goal.objects.filter(
            user=request.user, created_at__date__gte=from_date, created_at__date__lte=to_date
        )
        for goal in goals:
            tasks_linked = Task.objects.filter(goal=goal).count()
            habits_linked = Habit.objects.filter(goal=goal).count()
            ws.append([
                goal.title,
                goal.description or '',
                goal.status,
                round(goal.progress_percentage, 1),
                tasks_linked,
                habits_linked,
                goal.created_at.strftime('%Y-%m-%d')
            ])

    # REFLECTIONS SHEET
    if 'reflections' in include:
        ws = wb.create_sheet('Reflections')
        style_header(ws, ['Date', 'Wins', 'Challenges', "Tomorrow's Plan"])
        reflections = Reflection.objects.filter(
            user=request.user, date__gte=from_date, date__lte=to_date
        )
        for r in reflections:
            ws.append([
                str(r.date),
                r.content.get('wins', ''),
                r.content.get('challenges', ''),
                r.content.get('tomorrow', '')
            ])

    # Write to buffer and return as xlsx response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'LifeOS_Report_{from_date}_to_{to_date}.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

# ==================== ADMIN DASHBOARD ====================

def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_superuser:
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required
@admin_required
def admin_dashboard_overview(request):
    today = datetime.date.today()
    seven_days_ago = today - datetime.timedelta(days=7)

    total_users = User.objects.filter(is_superuser=False).count()

    active_users = User.objects.filter(
        is_superuser=False,
        is_active=True,
    ).filter(
        Q(task__created_at__date__gte=seven_days_ago) |
        Q(habit__habitcompletion__date__gte=seven_days_ago) |
        Q(reflection__date__gte=seven_days_ago)
    ).distinct().count()

    total_goals = Goal.objects.filter(user__is_superuser=False).count()
    total_tasks = Task.objects.filter(user__is_superuser=False).count()
    total_habits = Habit.objects.filter(user__is_superuser=False).count()

    all_completions = HabitCompletion.objects.filter(habit__user__is_superuser=False)
    if all_completions.exists():
        habit_completion_rate = all_completions.aggregate(avg=Avg('completion_percentage'))['avg']
    else:
        habit_completion_rate = 0

    context = {
        'total_users': total_users,
        'active_users': active_users,
        'total_goals': total_goals,
        'total_tasks': total_tasks,
        'total_habits': total_habits,
        'habit_completion_rate': round(habit_completion_rate),
    }
    return render(request, 'admin_overview.html', context)

@login_required
@admin_required
def admin_users(request):
    users = User.objects.filter(is_superuser=False).order_by('-date_joined')
    
    search_query = request.GET.get('q', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    filter_status = request.GET.get('status', 'all')
    if filter_status == 'active':
        users = users.filter(is_active=True)
    elif filter_status == 'inactive':
        users = users.filter(is_active=False)
        
    context = {
        'users': users,
        'search_query': search_query,
        'filter_status': filter_status,
    }
    return render(request, 'admin_users.html', context)

@login_required
@admin_required
def toggle_user_status(request, user_id):
    user = get_object_or_404(User, id=user_id, is_superuser=False)
    user.is_active = not user.is_active
    user.save()
    return redirect('admin_users')

@login_required
@admin_required
def admin_delete_user(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id, is_superuser=False)
        user.delete()
    return redirect('admin_users')

@login_required
@admin_required
def admin_activity(request):
    today = datetime.date.today()
    activity_data = []

    for i in range(29, -1, -1):
        day = today - datetime.timedelta(days=i)

        tasks_completed = Task.objects.filter(
            user__is_superuser=False,
            status='Completed',
            created_at__date=day
        ).count()

        habits_logged = HabitCompletion.objects.filter(
            habit__user__is_superuser=False,
            date=day,
            completion_percentage__gt=0
        ).count()

        activity_data.append({
            'date': day.strftime('%b %d'),
            'score': tasks_completed + habits_logged,
            'tasks': tasks_completed,
            'habits': habits_logged
        })

    context = {
        'activity_data_json': activity_data,
        'total_activity_period': sum(item['score'] for item in activity_data)
    }
    return render(request, 'admin_activity.html', context)

@login_required
@admin_required
def admin_habits(request):
    top_habits = Habit.objects.filter(
        user__is_superuser=False
    ).values('habit_name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]

    top_users = User.objects.filter(
        is_superuser=False
    ).annotate(
        log_count=Count('habit__habitcompletion')
    ).order_by('-log_count')[:5]

    context = {
        'top_habits': top_habits,
        'top_users': top_users,
    }
    return render(request, 'admin_habits.html', context)

@login_required
@admin_required
def admin_leaderboard(request):
    users = User.objects.filter(is_superuser=False)
    leaderboard = []

    def get_last_n_active_days(user, n):
        task_dates = list(Task.objects.filter(
            user=user, status='Completed'
        ).dates('created_at', 'day'))
        
        habit_dates = list(HabitCompletion.objects.filter(
            habit__user=user, completion_percentage__gt=0
        ).values_list('date', flat=True).distinct())
        
        all_dates = sorted(list(set(task_dates + habit_dates)), reverse=True)
        return all_dates[:n]

    for user in users:
        task_days = Task.objects.filter(
            user=user,
            status='Completed'
        ).dates('created_at', 'day').count()

        habit_days = HabitCompletion.objects.filter(
            habit__user=user,
            completion_percentage__gt=0
        ).values('date').distinct().count()

        active_days = max(task_days, habit_days)

        if active_days < 5:
            continue
            
        scores = []
        for day in get_last_n_active_days(user, active_days):
            habitScore = 0
            hab_comps = HabitCompletion.objects.filter(habit__user=user, date=day)
            h_count = hab_comps.count()
            if h_count > 0:
                habitScore = sum(hc.completion_percentage for hc in hab_comps) / h_count

            # FIX 2: use due_date-matched tasks as denominator (not all tasks ever created)
            taskScore = 0
            tasks_due = Task.objects.filter(user=user, due_date=day)
            t_total = tasks_due.count()
            if t_total > 0:
                t_completed = tasks_due.filter(status='Completed').count()
                taskScore = (t_completed / t_total) * 100

            scores.append((habitScore + taskScore) / 2)

        avg_score = sum(scores) / len(scores) if scores else 0

        leaderboard.append({
            'user': user,
            'avg_score': round(avg_score, 1),
            'active_days': active_days
        })

    leaderboard.sort(key=lambda x: x['avg_score'], reverse=True)
    
    context = {
        'leaderboard': leaderboard
    }
    return render(request, 'admin_leaderboard.html', context)

@login_required
@admin_required
def admin_age(request):
    age_groups = {
        '15-20': User.objects.filter(
            is_superuser=False,
            userprofile__age__gte=15, userprofile__age__lte=20
        ).count(),
        '21-25': User.objects.filter(
            is_superuser=False,
            userprofile__age__gte=21, userprofile__age__lte=25
        ).count(),
        '26-30': User.objects.filter(
            is_superuser=False,
            userprofile__age__gte=26, userprofile__age__lte=30
        ).count(),
        '30+': User.objects.filter(
            is_superuser=False,
            userprofile__age__gt=30
        ).count(),
    }
    
    def get_top_habits_by_age(min_age, max_age):
        if max_age:
            users = User.objects.filter(
                is_superuser=False,
                userprofile__age__gte=min_age,
                userprofile__age__lte=max_age
            )
        else:
            users = User.objects.filter(
                is_superuser=False,
                userprofile__age__gt=min_age
            )

        top_habits = Habit.objects.filter(
            user__in=users
        ).values('habit_name').annotate(
            count=Count('id')
        ).order_by('-count')[:5]

        return list(top_habits)
        
    context = {
        'count_15_20': age_groups['15-20'],
        'count_21_25': age_groups['21-25'],
        'count_26_30': age_groups['26-30'],
        'count_30_plus': age_groups['30+'],
        'habits_15_20': get_top_habits_by_age(15, 20),
        'habits_21_25': get_top_habits_by_age(21, 25),
        'habits_26_30': get_top_habits_by_age(26, 30),
        'habits_30_plus': get_top_habits_by_age(30, None),
    }
    return render(request, 'admin_age.html', context)
